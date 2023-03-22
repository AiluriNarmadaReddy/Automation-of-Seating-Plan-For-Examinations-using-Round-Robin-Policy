import time
import os.path
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.dataframe import dataframe_to_rows
import dataframes as di
import hall_plan as hz
import openpyxl
import out as ou

from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle, Color

# Get the current year, month, and date
year = time.strftime("%y", time.localtime())
mon = time.strftime("%b", time.localtime())
date = time.strftime("%d", time.localtime())

# Iterate through the halls and create a sheet for each hall
data = hz.hall_list
for i in range(len(data)):
    df = data[i]

sheet_names = []
for i in range(0, hz.no_of_halls):
    sheet_names.append(di.hall_df.iloc[i, 0])
    sheet_names = list(sheet_names)
print(sheet_names)

# Create the filename for the Excel file
excel_file_name = mon + "_" + year + ".xlsx"
print(excel_file_name)

# Check if the Excel file exists
if not os.path.exists(excel_file_name):
    # If the file doesn't exist, create a new workbook and sheet
    workbook = openpyxl.Workbook()
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    for i in range(0, len(sheet_names)):
            worksheet = workbook.create_sheet(str(sheet_names[i]))

            black = "FF000000"
            yellow = "FFFFFF00"
            thin = Side(style="thin", color=black)
            alignment = Alignment(horizontal="center", vertical="center")
            border_style = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill_style = PatternFill(fill_type="solid", patternType='solid', start_color=yellow)

            worksheet.merge_cells('A2:I2')
            worksheet['A2'].font = Font(size=24, name='Times New Roman')
            worksheet['A2'].alignment = alignment
            worksheet['A2'] = 'Mahaveer Institute of Science and Technology'
            worksheet['A2'].fill = fill_style
            worksheet['A2'].border = border_style

            worksheet.merge_cells('A3:I3')
            worksheet['A3'].font = Font(size=24, name='Times New Roman')
            worksheet['A3'].alignment = alignment
            title_string = 'Seating Arrangement' + " " + str(di.year[0]) + "-" + str(
                di.sem[0]) + " " + 'External Examination  ' + mon + "  " + year
            worksheet['A3'] = title_string
            worksheet['A3'].fill = fill_style
            worksheet['A3'].border = border_style

            worksheet.merge_cells('A4:C4')
            worksheet['A4'].font = Font(size=24, name='Times New Roman')
            worksheet['A4'].alignment = alignment
            worksheet['A4'] = "ROOM NO:" + str(sheet_names[i])
            worksheet.merge_cells('D4:E4')
            worksheet.merge_cells('F4:I4')
            worksheet['F4'].font = Font(size=24, name='Times New Roman')
            worksheet['F4'].alignment = alignment
            worksheet['F4'] = "DT:" + date + "." + mon + "." + year
            worksheet.merge_cells('A5:I5')
            output_str = "Strength:\n" + "\n".join([f"{k.title()}:{v}" for k, v in ou.output_dict[sheet_names[i]].items()])
            worksheet['A5'] = output_str
            worksheet['A5'].font = Font(size=24, name='Times New Roman')
            worksheet['A8'] = 'HT-NO'
            worksheet['A9'] = 'HT-NO'
            worksheet['A10'] = 'HT-NO'
            worksheet['A11'] = 'HT-NO'
            worksheet['A12'] = 'HT-NO'
            worksheet['A13'] = 'HT-NO'
            worksheet['A7'] = 'ENTRANCE'
            worksheet.column_dimensions['A'].auto_size = 'True'
            worksheet.merge_cells('B6:C6')
            worksheet['B6'] = 'ROW 1'
            worksheet['B6'].font = Font(size=24, name='Times New Roman')
            worksheet['B6'].alignment = alignment
            worksheet.merge_cells('D6:E6')
            worksheet['D6'] = 'ROW 2'
            worksheet['D6'].font = Font(size=24, name='Times New Roman')
            worksheet['D6'].alignment = alignment
            worksheet.merge_cells('F6:G6')
            worksheet['F6'] = 'ROW 3'
            worksheet['F6'].font = Font(size=24, name='Times New Roman')
            worksheet['F6'].alignment = alignment
            worksheet.merge_cells('H6:I6')
            worksheet['H6'] = 'ROW 4'
            worksheet['H6'].font = Font(size=24, name='Times New Roman')
            worksheet['H6'].alignment = alignment
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 15
            rows = range(5,14)

            # Set the height for each row
            height = 50
            for row in rows:
                worksheet.row_dimensions[row].height = height

            worksheet.column_dimensions['A'].width = 10
            i = sheet_names.index(sheet_names[i])

            for r in dataframe_to_rows(data[i], index=False, header=hz.headers[i]):
                worksheet.append(r)

            data_range = 'A2:I13'
            for row in worksheet[data_range]:
                for cell in row:
                    cell.border = cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = alignment
            source_range = worksheet['A14:H20']
            source_values = []
            for row in source_range:
                row_values = []
                for cell in row:
                    row_values.append(cell.value)
                source_values.append(row_values)

            # Get the target range (B7:I13)
            target_range = worksheet['B7:I13']

            # Write the values from the source range to the target range
            for i in range(len(source_values)):
                for j in range(len(source_values[i])):
                    target_range[i][j].value = source_values[i][j]
            for row in worksheet.iter_rows(min_row=14, max_row=20, min_col=1, max_col=8):
                for cell in row:
                    cell.value = None
            worksheet['A5'].alignment = Alignment(horizontal='right',vertical='center',wrap_text='True')
            row_num = 5
            max_height = 0
            for cell in worksheet[row_num]:
                if cell.value is not None:
                    lines = str(cell.value).split("\n")
                    cell_height = sum([len(line) // 3 * 14.4 + 10.0 for line in lines])
                    max_height = max(max_height, cell_height)
            worksheet.row_dimensions[row_num].height = max_height
    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size=4, orientation='landscape')

else:
# If the file exists, open it and get the active sheets
    workbook = openpyxl.load_workbook(excel_file_name)

workbook.save(excel_file_name)
workbook.close()
os.startfile(excel_file_name)
